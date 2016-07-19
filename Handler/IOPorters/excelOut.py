from openpyxl import Workbook


class ExcelOut:

    def __init__(self, ListValues, outputPath):
        self.path = outputPath
        self.listValues = ListValues




    def writer(self):
        wb = Workbook(guess_types=True)
        ws = wb.active
        col = 1
        row = 1
        lst = self.listValues
        firstRow = ["Component:", "Variable:", "Value:", "Unit:"]

        for stringa in firstRow:

            _ = ws.cell(column=col, row=row, value="%s" % stringa)
            col += 1

        row += 2
        col = 1


        for component in lst:

            _ = ws.cell(column=1, row=row, value="%s" % component)
            _ = ws.cell(column=2, row=row, value="Inputs:")

            row += 2


            for variable in lst[component]["Inputs"]:

                _ = ws.cell(column=2, row=row, value="%s" % variable)

                value = lst[component]["Inputs"][variable]["value"]
                if value is None:
                    pass
                elif type(value) is float:
                    fill = ws.cell(column=3, row=row, value="%g" % value)
                    fill.number_format = '0.00'
                else:
                    _ = ws.cell(column=3, row=row, value="%s" % value)

                unit = lst[component]["Inputs"][variable]["unit"]
                _ = ws.cell(column=4, row=row, value="%s" % unit)

                row += 1
            _ = ws.cell(column=2, row=row, value="EOC")
            row += 2

            _ = ws.cell(column=2, row=row, value="Attributes:")
            row += 2

            for variable in lst[component]["Attributes"]:

                _ = ws.cell(column=2, row=row, value="%s" % variable)

                value = lst[component]["Attributes"][variable]["value"]
                if value is None:
                    pass
                elif type(value) is float:
                    fill = ws.cell(column=3, row=row, value="%g" % value)
                    fill.number_format = '0.00'
                else:
                    _ = ws.cell(column=3, row=row, value="%s" % value)

                unit = lst[component]["Attributes"][variable]["unit"]
                _ = ws.cell(column=4, row=row, value="%s" % unit)

                row += 1
            row += 1



        _ = ws.cell(column=1, row=row, value="EOF")
        wb.save(filename=self.path)

        return "Output file correctly generated"
